from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "2024年工资汇总"

# 样式定义
header_font = Font(bold=True, size=12)
header_fill = PatternFill("solid", fgColor="4472C4")
header_font_white = Font(bold=True, size=12, color="FFFFFF")
title_font = Font(bold=True, size=14)
center = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题
ws.merge_cells('A1:H1')
ws['A1'] = "湛江市第七中学 - 黄燕妮老师 2024年度工资明细"
ws['A1'].font = Font(bold=True, size=16)
ws['A1'].alignment = center

# 人员基本信息
ws['A3'] = "姓名"
ws['B3'] = "黄燕妮"
ws['A4'] = "现任岗位/职称"
ws['B4'] = "中学二级教师"
ws['A5'] = "任职时间"
ws['B5'] = "2018年9月"
ws['A6'] = "人员类别"
ws['B6'] = "中职中小"
ws['A7'] = "学校"
ws['B7'] = "湛江市第七中学"

for row in range(3, 8):
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill("solid", fgColor="D9E1F2")

# 月度工资明细表头
headers = ["月份", "岗位工资", "薪级工资", "教护基本", "绩效工资", "应发合计", "扣款合计", "实发工资"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=10, column=col)
    cell.value = header
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# 月度数据（根据截图数据）
monthly_data = [
    ["1月", 1182, 380, 1551, 2425, 6798, 1525, 5273],
    ["2月", 1182, 380, 1551, 2425, 6798, 1525, 5273],
    ["3月", 1182, 380, 1551, 2425, 6798, 1525, 5273],
    ["4月", 1182, 380, 1551, 2425, 6798, 1525, 5273],
]

for row_idx, row_data in enumerate(monthly_data, 11):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = center
        if col_idx > 1:
            cell.number_format = '#,##0'

# 汇总行
ws['A15'] = "年度合计"
ws['A15'].font = Font(bold=True)
ws['A15'].fill = PatternFill("solid", fgColor="FFC000")
ws['A15'].alignment = center

for col in range(2, 9):
    cell = ws.cell(row=15, column=col)
    col_letter = get_column_letter(col)
    cell.value = f"=SUM({col_letter}11:{col_letter}14)"
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="FFC000")
    cell.border = thin_border
    cell.alignment = center
    if col > 1:
        cell.number_format = '#,##0'

# 平均行
ws['A16'] = "月均"
ws['A16'].font = Font(bold=True)
ws['A16'].fill = PatternFill("solid", fgColor="92D050")
ws['A16'].alignment = center

for col in range(2, 9):
    cell = ws.cell(row=16, column=col)
    col_letter = get_column_letter(col)
    cell.value = f"=AVERAGE({col_letter}11:{col_letter}14)"
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="92D050")
    cell.border = thin_border
    cell.alignment = center
    if col > 1:
        cell.number_format = '#,##0'

# 设置列宽
widths = [12, 14, 14, 14, 14, 14, 14, 14]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 保存
wb.save(r"C:\Users\南\Desktop\工资汇总表_黄燕妮_2024.xlsx")
print("Excel文件已创建成功！")